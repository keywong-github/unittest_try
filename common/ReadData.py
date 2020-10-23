import openpyxl

class ReadData1:
    def get_excel_data(self):
        excel1=openpyxl.load_workbook('D:\\2020\\vscode\mine\my-autotest\common\data.xlsx')
        sheet1=excel1['login']

        max_row=sheet1.max_row
        max_column=sheet1.max_column
        listdata=[]

        for x in range(2,max_row+1):
            dict={}
            for y in range(1,max_column+1):
                value=sheet1.cell(x,y).value
                dict[sheet1.cell(1,y).value]=value
            listdata.append(dict)
        #print(listdata)
        
        return listdata

#ReadData1().get_excel_data()
